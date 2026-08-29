#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
泰安IDC配比出口流量分析工具
功能：自动从高到低排列流量数据，并标红95th百分位数据
"""

import csv
import sys
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("正在安装openpyxl库...")
    os.system("pip3 install openpyxl -q")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

import numpy as np


def read_csv_data(filepath):
    """读取CSV文件并解析流量数据"""
    data = []
    
    with open(filepath, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        
        # 读取标题行
        title_row = next(reader)
        title = title_row[1] if len(title_row) > 1 else "未知"
        
        # 读取元数据行
        metadata = {}
        p95_value = None
        for i in range(8):
            row = next(reader)
            if len(row) >= 2:
                metadata[row[0]] = row[1]
                # 第9行包含95th值: "百分之百","26249485772.8","|95:bits:0:aggregate_peak:2|"
                if row[0] == '百分之百' and len(row) >= 3:
                    try:
                        p95_value = float(row[1])
                    except:
                        pass
        
        # 读取表头
        header = next(reader)
        
        # 读取数据行
        for row in reader:
            if len(row) >= 9 and row[0].startswith('2026'):
                try:
                    date = row[0]
                    total1 = float(row[6]) if row[6] else 0
                    total2 = float(row[8]) if row[8] else 0
                    
                    # 过滤无效数据
                    if total1 > 0 and total2 > 0:
                        # 取两组数据的较大值作为计算依据
                        max_value = max(total1, total2)
                        data.append({
                            'date': date,
                            'total1': total1,
                            'total2': total2,
                            'total': max_value
                        })
                except:
                    continue
    
    return title, metadata, data, p95_value


def get_month_days(start_date, end_date):
    """自动判断月份天数"""
    from datetime import datetime
    
    start = datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S')
    end = datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S')
    
    # 计算月份天数
    if start.month == end.month:
        # 同一个月
        import calendar
        days = calendar.monthrange(start.year, start.month)[1]
    else:
        # 跨月（取开始月的天数）
        import calendar
        days = calendar.monthrange(start.year, start.month)[1]
    
    return days


def calculate_statistics(data):
    """计算统计数据"""
    totals = np.array([d['total'] for d in data])
    
    stats = {
        'count': len(data),
        'max': totals.max(),
        'min': totals.min(),
        'mean': totals.mean(),
        'median': np.median(totals),
        'p95': np.percentile(totals, 95),
        'std': totals.std()
    }
    
    return stats


def calculate_monthly_95th(data):
    """按整月计算95th值"""
    # 获取时间范围
    start_date = data[-1]['date']
    end_date = data[0]['date']
    
    # 计算月份天数
    month_days = get_month_days(start_date, end_date)
    
    # 理论采样点数（每5分钟一个采样，每天288个）
    samples_per_day = 288
    total_samples_theory = month_days * samples_per_day
    
    # 实际采样点数
    actual_samples = len(data)
    
    # 按理论采样点数计算95th位置（去掉最高的5%）
    p95_rank_theory = int(total_samples_theory * 0.05) + 1  # 第几名
    
    # 按实际数据排序
    sorted_totals = sorted([d['total'] for d in data], reverse=True)
    
    # 如果实际数据不足理论数量，用实际数据的95th位置
    # 如果实际数据足够，用理论位置
    if actual_samples >= p95_rank_theory:
        p95_index = p95_rank_theory - 1  # 数组索引从0开始
    else:
        # 实际数据不足，按实际数据计算
        p95_index = int(actual_samples * 0.05)
    
    p95_value = sorted_totals[p95_index]
    p95_rank = p95_index + 1  # 实际排名
    
    return {
        'month_days': month_days,
        'samples_per_day': samples_per_day,
        'total_samples_theory': total_samples_theory,
        'actual_samples': actual_samples,
        'p95_rank_theory': p95_rank_theory,  # 理论95th位置
        'p95_rank': p95_rank,  # 实际95th位置
        'p95_value': p95_value
    }


def sort_data_by_traffic(data):
    """按流量从高到低排序"""
    return sorted(data, key=lambda x: x['total'], reverse=True)


def get_nominal_95th(title):
    """从标题中提取标称95th带宽"""
    # 尝试从标题中提取，如 "泰安IDC配比出口-42G"
    if '-' in title:
        bandwidth = title.split('-')[-1]
        try:
            return float(bandwidth.replace('G', '').replace('g', '')) * 1e9
        except:
            pass
    return None


def create_excel_report(sorted_data, stats, title, metadata, output_path, monthly_info, current_p95, monthly_p95):
    """创建Excel报告"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "流量分析"
    
    # 定义样式
    header_font = Font(name='微软雅黑', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 95th数据样式（红色背景）
    p95_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    p95_font = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
    
    # 普通数据样式
    normal_font = Font(name='微软雅黑', size=10)
    normal_alignment = Alignment(horizontal='center', vertical='center')
    
    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========== 标题区域 ==========
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"泰安IDC配比出口流量分析报告 - {title}"
    title_cell.font = Font(name='微软雅黑', bold=True, size=16, color='2F5496')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40
    
    # ========== 95th核心指标 ==========
    ws.merge_cells('A3:H3')
    ws['A3'].value = "95th计费指标"
    ws['A3'].font = Font(name='微软雅黑', bold=True, size=12, color='2F5496')
    ws['A3'].fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    
    # 计算月95th值（基于整月理论采样点）
    monthly_95_value = monthly_p95['p95_value']
    
    p95_data = [
        ['当前95th值', f"{current_p95/1e9:.2f} Gbps", f"第{monthly_info['p95_rank']}名"],
        ['月95th值', f"{monthly_95_value/1e9:.2f} Gbps", f"第{monthly_p95['p95_rank_theory']}名"],
        ['数据时间范围', f"{sorted_data[-1]['date']} 至 {sorted_data[0]['date']}"],
        ['有效数据点', f"{stats['count']} 个"]
    ]
    
    for i, row_data in enumerate(p95_data, start=4):
        ws[f'A{i}'].value = row_data[0]
        ws[f'A{i}'].font = Font(name='微软雅黑', bold=True, size=10)
        ws[f'B{i}'].value = row_data[1]
        ws[f'B{i}'].font = normal_font
        if len(row_data) > 2:
            ws[f'C{i}'].value = row_data[2]
            ws[f'C{i}'].font = normal_font
    
    # ========== 数据表格 ==========
    start_row = 10
    
    # 表头
    headers = ['排名', '时间', '流量(Gbps)', '百分位', '备注']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 设置列宽
    column_widths = [8, 22, 15, 10, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # 填充数据
    total_samples = len(sorted_data)
    
    # 找到95th位置的数据点（从高到低第5%位置）
    p95_index = int(total_samples * 0.05)  # 第5%位置
    p95_rank = p95_index + 1
    
    for idx, record in enumerate(sorted_data):
        row = start_row + 1 + idx
        rank = idx + 1
        percentile = (total_samples - rank) / total_samples * 100
        is_p95_point = (rank == p95_rank)  # 只标红95th那一个点
        
        # 写入数据（只显示较大值）
        ws.cell(row=row, column=1, value=rank)
        ws.cell(row=row, column=2, value=record['date'])
        ws.cell(row=row, column=3, value=round(record['total']/1e9, 2))
        ws.cell(row=row, column=4, value=f"{percentile:.1f}%")
        
        # 备注
        if is_p95_point:
            ws.cell(row=row, column=5, value=f'★ 95th计费点 ({record["total"]/1e9:.2f}G)')
        elif record['total'] > 42e9:
            ws.cell(row=row, column=5, value='⚠️ 超过42G')
        else:
            ws.cell(row=row, column=5, value='')
        
        # 设置样式
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.alignment = normal_alignment
            cell.border = thin_border
            
            # 只标红95th那一个点
            if is_p95_point:
                cell.fill = p95_fill
                cell.font = p95_font
            else:
                cell.font = normal_font
    
    # 添加说明
    note_row = start_row + total_samples + 2
    ws.merge_cells(f'A{note_row}:E{note_row}')
    p95_value = sorted_data[p95_index]['total'] / 1e9
    ws[f'A{note_row}'].value = f"说明：红色背景为95th计费点（第{p95_rank}名，流量 {p95_value:.2f}G）"
    ws[f'A{note_row}'].font = Font(name='微软雅黑', italic=True, size=10, color='FF0000')
    
    # ========== 保存文件 ==========
    wb.save(output_path)
    print(f"✓ 报告已生成: {output_path}")


def main():
    """主函数"""
    print("=" * 60)
    print("泰安IDC配比出口流量分析工具")
    print("=" * 60)
    
    # 输入文件
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
    else:
        input_file = '/opt/excel/泰安IDC配比出口-42G.csv'
    
    # 输出文件
    output_file = input_file.replace('.csv', '_分析报告.xlsx')
    
    if not os.path.exists(input_file):
        print(f"错误: 找不到文件 {input_file}")
        sys.exit(1)
    
    print(f"\n读取文件: {input_file}")
    
    # 读取数据
    title, metadata, data, p95_from_csv = read_csv_data(input_file)
    print(f"✓ 成功读取 {len(data)} 条有效数据")
    
    # 计算统计
    stats = calculate_statistics(data)
    print(f"✓ 统计计算完成")
    
    # 按实际有效数据计算95th（当前95）
    total_samples = len(data)
    top_5_percent = int(total_samples * 0.05)
    p95_rank = top_5_percent + 1  # 第几名
    
    # 排序数据
    sorted_totals = sorted([d['total'] for d in data], reverse=True)
    current_p95 = sorted_totals[p95_rank - 1]  # 当前95th值
    
    # 计算月份信息（月95）
    monthly_info = calculate_monthly_95th(data)
    monthly_info['p95_rank'] = p95_rank
    monthly_info['p95_value'] = current_p95
    
    # 计算月95th值（基于整月理论采样点）
    monthly_p95_rank = monthly_info['p95_rank_theory']
    if len(sorted_totals) >= monthly_p95_rank:
        monthly_p95_value = sorted_totals[monthly_p95_rank - 1]
    else:
        monthly_p95_value = current_p95  # 数据不足时使用当前95
    
    monthly_p95 = {
        'p95_rank_theory': monthly_p95_rank,
        'p95_value': monthly_p95_value
    }
    
    print(f"\n【95th计算】")
    print(f"  - 有效数据点: {total_samples}")
    print(f"  - 当前95th值: {current_p95/1e9:.2f} Gbps (第{p95_rank}名)")
    print(f"  - 月95th值: {monthly_p95_value/1e9:.2f} Gbps (第{monthly_p95_rank}名)")
    
    # 排序数据
    sorted_data = sort_data_by_traffic(data)
    print(f"✓ 数据已按流量从高到低排序")
    
    # 生成报告
    create_excel_report(sorted_data, stats, title, metadata, output_file, monthly_info, current_p95, monthly_p95)
    
    # 统计高于95th的数据点
    above_p95_count = sum(1 for d in data if d['total'] > current_p95)
    
    print("\n" + "=" * 60)
    print("分析完成!")
    print("=" * 60)
    print(f"\n输出文件: {output_file}")
    print(f"\n主要结果:")
    print(f"  - 总数据点: {stats['count']}")
    print(f"  - 最大流量: {stats['max']/1e9:.2f} Gbps")
    print(f"  - 95th计费值: {current_p95/1e9:.2f} Gbps")
    print(f"  - 95th位置: 第{monthly_info['p95_rank']}名")
    print(f"  - 高于95th的数据点: {above_p95_count} ({above_p95_count/stats['count']*100:.1f}%)")
    print(f"  - 已标红第{monthly_info['p95_rank']}名")


if __name__ == '__main__':
    main()
